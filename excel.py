import os
import numpy as np
import math
import sys
import re
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from statistics import geometric_mean
from openpyxl.styles import PatternFill, Font, GradientFill, Alignment, Border, Side
from openpyxl.styles.colors import Color

bench = str(sys.argv[1])
overhead = str(sys.argv[2])
# path = str(sys.argv[3])
fName = bench + "-"+overhead
defaultSoftfileName = os.getcwd() + "/" + bench + "/" + \
    fName + "-default-out.txt"
changingSoftfileName = os.getcwd() + "/" + bench + "/" + fName + "-out.txt"
fixedSoftfileName = os.getcwd() + "/" + bench + "/" + \
    fName + "-adoptive-out.txt"

defaultSoftData = []
changingSoftData = []
fixedSoftData = []

normalizedChangingSoftHeap = []  # normalized to default
normalizedFixedSoftHeap = []  # normalized to default
normalizedChangingSoftHeap.append('Variable Soft Heap with /*2 Formula')
normalizedFixedSoftHeap.append('Variable Soft Heap with Adoptive Formula')
normalizedChangingSoftHeap.append(float(overhead))
normalizedFixedSoftHeap.append(float(overhead))

defaultSoftData.append('Default Soft Heap')
changingSoftData.append('Variable Soft Heap with /*2 Formula')
fixedSoftData.append('Variable Soft Heap with Adoptive Formula')
defaultSoftData.append(float(overhead))
changingSoftData.append(float(overhead))
fixedSoftData.append(float(overhead))

print ("****************", fName, "****************")
if (os.path.exists(defaultSoftfileName)):
    with open(defaultSoftfileName, 'r') as f:
        defaultSoftData.extend(line.strip() for line in f)
        print(defaultSoftData)

if (os.path.exists(changingSoftfileName)):
    with open(changingSoftfileName, 'r') as f:
        changingSoftData.extend(line.strip() for line in f)
        print(changingSoftData)

if (os.path.exists(fixedSoftfileName)):
    with open(fixedSoftfileName, 'r') as f:
        fixedSoftData.extend(line.strip() for line in f)
        print(fixedSoftData)

for i in range(len(defaultSoftData)):
    #print(changingSoftData[i])
    if (i > 1):
        if(float(defaultSoftData[i])>float(0)):
            normalizedChangingSoftHeap.append(
           "{:.2f}".format(float(changingSoftData[i])/float(defaultSoftData[i])))
            normalizedFixedSoftHeap.append(
            "{:.2f}".format(float(fixedSoftData[i])/float(defaultSoftData[i])))
        else:
            normalizedChangingSoftHeap.append(float(0))
            normalizedFixedSoftHeap.append(float(0))
curr = os.getcwd() + '/doc.xlsx'

if (os.path.exists(curr)):
    wb = load_workbook('doc.xlsx')
else:
    wb = Workbook()


if bench in wb.sheetnames:
    sheet = wb[bench]
else:
    sheet = wb.create_sheet(bench)
    sheet.merge_cells('A1:G1')
    sheet['A1'] = 'Normalized to Default Heap Size Behavior'
    # header = (['', 'GC overhead', 'GC calls', '90th Per - execution times',
    #            '90th Per - pause times GC', '90th per - heap before GC', '90th per - heap after GC'])
    header = (['', 'GC overhead', 'GC calls', '90th Per - execution times',
               '90th Per - 90th per latencies', '90th per - heap before GC', '90th per - heap after GC'])
    sheet.append(header)

# sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

empty_row = (['', '', '', '', '', ''])

style = PatternFill(start_color='00FFCC99', end_color='00FFCC99',
                    fill_type="solid")


# sheet['A1'].fill = style

leng = len('90th Per - pause times GC')
sheet.column_dimensions['A'].width = len(
    'Variable Soft Heap with Adoptive Formula')
sheet.column_dimensions['B'].width = leng
sheet.column_dimensions['C'].width = leng
sheet.column_dimensions['D'].width = leng
sheet.column_dimensions['E'].width = leng
sheet.column_dimensions['F'].width = leng
sheet.column_dimensions['G'].width = leng


for rows in sheet.iter_rows(min_row=2, min_col=1, max_row=2, max_col=7):
    for cell in rows:
        cell.fill = style
# for rows in sheet.iter_rows(min_row=6, min_col=1, max_row=6, max_col=7):
#     for cell in rows:
#         cell.fill = style
# sheet.append(defaultSoftData)
# sheet.append(changingSoftData)
# sheet.append(fixedSoftData)
# sheet.append(empty_row)
sheet.append(normalizedChangingSoftHeap)
sheet.append(normalizedFixedSoftHeap)
sheet.append(empty_row)
r = sheet.max_row
c = sheet.max_column
# for i in range(r+1, r+1):
#     for j in range(1, c+1):
#         cel = sheet.cell(row=r , col=c, value = '')

wb.save('doc.xlsx')
